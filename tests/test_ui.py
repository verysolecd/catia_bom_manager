import sys
import openpyxl
from PyQt5.QtWidgets import QApplication, QTableView, QWidget, QVBoxLayout
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtCore import Qt, QSortFilterProxyModel
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import QShortcut


# 创建 Excel 文件并写入数据
def create_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in range(1, 11):
        for col in range(1, 11):
            sheet.cell(row=row, column=col, value=f"Cell {row - 1}-{col - 1}")
    workbook.save("example.xlsx")


class CustomTableView(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self)
        self.copy_shortcut.activated.connect(self.copy_cells)
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self)
        self.paste_shortcut.activated.connect(self.paste_cells)

    def copy_cells(self):
        selected_indexes = self.selectedIndexes()
        if selected_indexes:
            copy_text = ""
            rows = sorted(set(index.row() for index in selected_indexes))
            cols = sorted(set(index.column() for index in selected_indexes))
            for row in rows:
                for col in cols:
                    index = self.model().index(row, col)
                    if index in selected_indexes:
                        copy_text += str(index.data())
                    if col < cols[-1]:
                        copy_text += "\t"
                copy_text += "\n"
            clipboard = QApplication.clipboard()
            clipboard.setText(copy_text)

    def paste_cells(self):
        clipboard = QApplication.clipboard()
        paste_text = clipboard.text()
        rows = paste_text.split('\n')
        selected_indexes = self.selectedIndexes()
        if selected_indexes:
            top_left = selected_indexes[0]
            for i, row in enumerate(rows):
                cols = row.split('\t')
                for j, col in enumerate(cols):
                    index = self.model().index(top_left.row() + i, top_left.column() + j)
                    if index.isValid():
                        self.model().setData(index, col)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()

        # 创建 Excel 文件
        create_excel()

        # 读取 Excel 文件
        workbook = openpyxl.load_workbook("example.xlsx")
        sheet = workbook.active

        # 定义要截取的区域（这里截取第 2 行到第 6 行，第 2 列到第 6 列）
        start_row = 2
        end_row = 6
        start_col = 2
        end_col = 6

        # 创建模型
        self.model = QStandardItemModel(end_row - start_row + 1, end_col - start_col + 1)
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                item = QStandardItem(str(sheet.cell(row=row, column=col).value))
                self.model.setItem(row - start_row, col - start_col, item)

        # 创建自定义表格视图
        self.table_view = CustomTableView()
        self.table_view.setModel(self.model)

        layout.addWidget(self.table_view)
        self.setLayout(layout)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

