import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QMessageBox
import openpyxl
import subprocess


# 主窗口模块
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # 创建按钮
        self.button = QPushButton('打开 Excel 并注入数据', self)
        self.button.clicked.connect(self.open_excel_and_inject_data)

        # 布局
        layout = QVBoxLayout()
        layout.addWidget(self.button)
        self.setLayout(layout)

        self.setWindowTitle('Excel 操作示例')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def open_excel_and_inject_data(self):
        # 禁用按钮和窗口
        self.button.setEnabled(False)
        self.setDisabled(True)

        # 调用 Excel 操作模块的函数
        excel_operator = ExcelOperator()
        excel_operator.open_excel()
        excel_operator.inject_data()

        # 打开 Excel 文件供用户操作
        try:
            file_path = os.path.join('resource', 'injdata.xlsx')
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            else:  # Linux 和 macOS
                subprocess.call(('open', file_path))
        except Exception as e:
            QMessageBox.critical(self, '错误', f'打开 Excel 文件时出错: {str(e)}')
            self.button.setEnabled(True)
            self.setDisabled(False)
            return

        # 提示用户手动操作 Excel 并保存
        QMessageBox.information(
            self, '提示', '请手动操作 Excel 文件，操作完成后保存并关闭文件，然后点击确定继续...')

        # 读取用户保存后的 Excel 数据
        data = excel_operator.read_data_to_array()
        # 调用数据处理模块的函数
        data_processor = DataProcessor()
        data_processor.process_data(data)

        # 启用按钮和窗口
        self.button.setEnabled(True)
        self.setDisabled(False)


# Excel 操作模块
class ExcelOperator:
    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def open_excel(self):
        file_path = os.path.join('resources', 'injdata.xlsx')
        # 检查 resource 文件夹是否存在，不存在则创建
        if not os.path.exists('resources'):
            os.makedirs('resources')
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def inject_data(self):
        # 向 Excel 中注入数据
        data_to_inject = [
            ['姓名', '年龄', '性别'],
            ['张三', 25, '男'],
            ['李四', 30, '女']
        ]
        for row in data_to_inject:
            self.worksheet.append(row)
        file_path = os.path.join('resource', 'injdata.xls')
        self.workbook.save(file_path)

    def read_data_to_array(self):
        file_path = os.path.join('resource', 'injdata.xls')
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.worksheet = self.workbook.active
            data = []
            for row in self.worksheet.iter_rows(values_only=True):
                data.append(list(row))
            return data
        except Exception as e:
            print(f'读取 Excel 文件时出错: {str(e)}')
            return []


# 数据处理模块
class DataProcessor:
    def process_data(self, data):
        # 对从 Excel 中提取的数据进行处理
        print("从 Excel 中提取的数据：")
        for row in data:
            print(row)
        # 可以在这里添加更多的数据处理逻辑，例如提取特定列的数据到数组


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec_())
